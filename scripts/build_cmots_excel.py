"""Build the CMOTS-equivalent API catalog Excel workbook.

Output: data/exports/CMOTS_Data_Catalog.xlsx with 3 sheets:
  1. EQUITY          - ~95 equity APIs, one row per output field
  2. MUTUAL FUNDS    - ~40 MF APIs, one row per output field
  3. CMOTS COMMENTS  - Implementation remarks
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from scripts.cmots.equity_part1 import EQUITY_APIS_1_30
from scripts.cmots.equity_part2 import EQUITY_APIS_31_65
from scripts.cmots.equity_part3 import EQUITY_APIS_66_100
from scripts.cmots.mutual_funds import MF_APIS
from scripts.cmots.comments import CMOTS_COMMENTS

EQUITY_APIS = EQUITY_APIS_1_30 + EQUITY_APIS_31_65 + EQUITY_APIS_66_100

HEADER_COLS = [
    "API Number", "Report Name", "API URL", "Frequency", "Updation Time",
    "Input", "Input Description", "Output", "Data Type", "Output Description",
]

# Audit columns appended to every API's outputs (per automation spec).
AUDIT_OUTPUTS = [
    ("source_name", "TEXT", "Origin source identifier (nse/bse/screener/amfi/moneycontrol)"),
    ("last_refresh_time", "TIMESTAMPTZ", "When this row was last refreshed from source"),
    ("data_version", "INTEGER", "Auto-incremented version number for change tracking"),
    ("created_at", "TIMESTAMPTZ", "Row creation timestamp (UTC)"),
    ("updated_at", "TIMESTAMPTZ", "Row last-update timestamp (UTC)"),
]

# ─── Styling ──────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

ALT_FILL_A = PatternFill("solid", fgColor="FFFFFF")
ALT_FILL_B = PatternFill("solid", fgColor="F2F7FB")

BODY_FONT = Font(name="Calibri", size=10)
BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN = Side(border_style="thin", color="C0C0C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_WIDTHS_API = {
    "API Number": 12, "Report Name": 32, "API URL": 42, "Frequency": 14,
    "Updation Time": 22, "Input": 22, "Input Description": 36,
    "Output": 28, "Data Type": 18, "Output Description": 50,
}


def format_inputs(inputs: list[tuple[str, str]]) -> tuple[str, str]:
    """Concatenate inputs into two parallel multi-line strings."""
    names = "\n".join(name for name, _ in inputs) if inputs else ""
    descs = "\n".join(desc for _, desc in inputs) if inputs else ""
    return names, descs


def write_api_sheet(ws, apis: list[dict]) -> int:
    """Write APIs to sheet, one row per output field. Returns row count."""
    # Header
    for col_idx, col_name in enumerate(HEADER_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER

    row = 2
    for api_idx, api in enumerate(apis):
        in_names, in_descs = format_inputs(api.get("inputs", []))
        api_metadata = [
            api["num"], api["name"], api["url"], api["freq"], api["update"],
            in_names, in_descs,
        ]
        fill = ALT_FILL_A if api_idx % 2 == 0 else ALT_FILL_B

        # Append audit columns to every API's output list
        outputs = list(api.get("outputs", []) or [("", "", "")]) + AUDIT_OUTPUTS
        for out_name, out_type, out_desc in outputs:
            values = api_metadata + [out_name, out_type, out_desc]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = BODY_FONT
                cell.alignment = BODY_ALIGN
                cell.fill = fill
                cell.border = BORDER
            row += 1

    # Column widths
    for col_idx, col_name in enumerate(HEADER_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS_API[col_name]

    # Freeze header
    ws.freeze_panes = "A2"
    return row - 2


def write_comments_sheet(ws, comments: list[tuple[str, str]]) -> int:
    headers = ["Remark", "Cmots comments"]
    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER

    for r_idx, (remark, comment) in enumerate(comments, start=2):
        fill = ALT_FILL_A if r_idx % 2 == 0 else ALT_FILL_B
        for c_idx, val in enumerate([remark, comment], start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.fill = fill
            cell.border = BORDER

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = "A2"
    return len(comments)


def build_workbook(output_path: Path) -> dict:
    wb = Workbook()
    ws_eq = wb.active
    ws_eq.title = "EQUITY"
    ws_mf = wb.create_sheet("MUTUAL FUNDS")
    ws_cm = wb.create_sheet("CMOTS COMMENTS")

    eq_rows = write_api_sheet(ws_eq, EQUITY_APIS)
    mf_rows = write_api_sheet(ws_mf, MF_APIS)
    cm_rows = write_comments_sheet(ws_cm, CMOTS_COMMENTS)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {
        "path": str(output_path),
        "equity_apis": len(EQUITY_APIS),
        "equity_rows": eq_rows,
        "mf_apis": len(MF_APIS),
        "mf_rows": mf_rows,
        "comment_rows": cm_rows,
    }


if __name__ == "__main__":
    out = ROOT / "data" / "exports" / "CMOTS_Data_Catalog.xlsx"
    try:
        result = build_workbook(out)
    except PermissionError:
        # File is open in Excel — write to a v2 file
        out = ROOT / "data" / "exports" / "CMOTS_Data_Catalog_v2.xlsx"
        print(f"  [PermissionError on primary file — writing to {out.name}]")
        result = build_workbook(out)
    print(f"\n  EQUITY sheet:        {result['equity_apis']} APIs, {result['equity_rows']} rows")
    print(f"  MUTUAL FUNDS sheet:  {result['mf_apis']} APIs, {result['mf_rows']} rows")
    print(f"  CMOTS COMMENTS:      {result['comment_rows']} rows")
    print(f"\n  Saved: {result['path']}")
    print(f"  Size:  {os.path.getsize(out)/1024:.1f} KB\n")
